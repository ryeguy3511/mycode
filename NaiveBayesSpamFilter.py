"""
Spam Filter
"""

from functools import reduce
import operator
from openpyxl import load_workbook
from textblob import TextBlob

class SpamClassifier():
    
    def fit(self, trainDataS, trainLabelS, trainDataH, trainLabelH):
        self.trainDataS = trainDataS
        self.trainLabelS = trainLabelS
        self.trainDataH = trainDataH
        self.trainLabelH = trainLabelH
        
    
    def predict(self, usrMessage):
        self.Prob(usrMessage)
               
    def Prob(self, usrMessage):
        # creating empty lists for individual word probabilites 
        global ps
        ps = []
        global psminus
        psminus = []
        global ph
        ph = []
        global phminus
        phminus = []
        for word in usrMessage:
            ps.append(trainDataS.count(word) / len(trainDataS)) 
            psminus.append(1-(trainDataS.count(word) / len(trainDataS)))
            
        for word in usrMessage:
            ph.append(trainDataH.count(word) / len(trainDataH))
            phminus.append(1-(trainDataH.count(word) / len(trainDataH)))
        
            global psmessage 
            psmessage = prod(ps) / (prod(ps) + prod(psminus))
            global phmessage 
            phmessage = prod(ph) / (prod(ph) + prod(phminus))
        
        print(f"Probability spam is: {psmessage*100} %")
        print(f"Probability ham is: {phmessage*100} %")
        
        if psmessage > phmessage:
            print("This message is spam")
        elif psmessage < phmessage:
            print("This message is ham")
        else:
            print("Inconclusive")
            
def prod(iterable):
    return reduce(operator.mul, iterable, 1)

# flatten function

def flatten(aList):
    new_list = []
    found_list = True
    while found_list:
        found_list = False
        for element in aList:
            if isinstance(element, list):
                for subelement in element:
                    new_list.append(subelement)
                found_list = True
            else:
                new_list.append(element)
            
        new_list, aList = aList, new_list
        new_list.clear()
    return aList

# tokenize words in message
    
def process_message(message):
    message = TextBlob(message)
    words = message.words.lower()
    return words


# load excel file

wb = load_workbook('spamfilterdata.xlsx')
sheetSpam = wb['spam_data']
sheetHam = wb['ham_data']
cellRange = len(sheetSpam[1:50]) + 1

trainDataS = []
trainDataH = []
trainLabelS = []
trainLabelH = []


# add the spam data to the training data in the form of independent words
for i in range(1, cellRange):
    words = process_message(sheetSpam.cell(row = i, column = 2).value)
    trainDataS.append(words)
    trainDataS = flatten(trainDataS)

lenSpam = len(trainDataS) + 1

for i in range(1, lenSpam):
    trainLabelS.append('spam')
    
for i in range(1, cellRange):
    words = process_message(sheetHam.cell(row = i, column = 2).value)
    trainDataH.append(words)
    trainDataH = flatten(trainDataH)

lenHam = len(trainDataH) + 1

for i in range(1, lenHam):
    trainLabelH.append('ham')
    
clf = SpamClassifier()
clf.fit(trainDataS, trainLabelS, trainDataH, trainLabelH)

usrMessage = input("Please input a subject header: ")
usrMessage = process_message(usrMessage)
clf.predict(usrMessage)


